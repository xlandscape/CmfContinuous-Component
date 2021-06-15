
import os
from shutil import copyfile
import pandas as pd
import numpy as np
import subprocess    
import datetime
import matplotlib
from pandas.tools.plotting import table
import matplotlib.lines as mlines
import matplotlib.patches as mpatches
from openpyxl import load_workbook
import matplotlib.pylab as plt


class MACRO_pestmod():
        
    def __init__(self,fname,fpath):
        # path of input parameter file
        self.fpath = fpath
        self.fname=fname
        # name of sheets in excelfile
        self.tablenames = ["MODELRUNS","HEADER","PHYSICAL","PROPERTIES","OPTIONS","SOLUTE","CROP","SITE","IRRIGATION","SETUP","TILLAGE","OUTPUTS","BOUNDARY AND INITIAL CONDITIONS"]
        self.headers_parfile = ["","PHYSICAL PARAMETERS","PROPERTIES","OPTIONS","SOLUTE PARAMETERS","CROP PARAMETERS","SITE PARAMETERS","IRRIGATION PARAMETERS","SETUP","TILLAGE PARAMETERS","OUTPUTS","BOUNDARY AND INITIAL CONDITIONS"]
        #open excel file and load data
        self.xls = pd.ExcelFile(os.path.join(self.fpath,self.fname + ".xlsx"))
        tables = [self.xls.parse(name) for name in self.tablenames]
        self.tables=dict(zip(self.tablenames,tables))
        self.helper_soil = self.xls.parse("helper_soil")
        self.helper_GAP = self.xls.parse("helper_GAP")
        self.helper_crop = self.xls.parse("helper_crop")
        
        
        self.modelruns_status = []
        
        #TODO

    def create_results_summary(self):
        """
        Creates summary file of all model runs.
        """
        
        # get mdoelruns
        modelruns = self.tables["MODELRUNS"]
        
        #new column names
        columns_new=["Date","Rain [mm/h]","Precip [mm/h]","Drainage [mm/h]",
                     "Runoff [mm/h]","Perc [mm/h]","Evaporation [mm/h]",
                     "Transpiration [mm/h]","Drainage sec. [mm/h]",
                     "Solutes perc accumulated [mg/m2]","Solutes runoff accumulated [mg/m2]",
                     "Solutes micropores liquid [mg/m2]","Solutes macropores liquid [mg/m2]",
                     "Solutes micropores solid [mg/m2]","Solutes macropores solid [mg/m2]",
                     "Solutes degraded accumulated [mg/m2]","Solutes drainage accumulated [mg/m2]",
                     "Solutes sec drainage accumulated [mg/m2]","Solutes gw losses accumulated [mg/m2]",
                     "ADMI_KIN"]  
        
        # new column data types
        dtype={'ADMI_KIN                      0-  0          001   0': "float32",
                 'DRAINLOS                                     001    ': "float32",
                 'Date': 'int64',
                 'ESOIL                                        001    ': "float32",
                 'ETA                                          001    ': "float32",
                 'GSINK                                        001    ': "float32",
                 'GSRAT                                        001    ': "float32",
                 'RAIN                                         001    ': "float32",
                 'SRUNOFF                                      001    ': "float32",
                 'SSEEP                                        001    ': "float32",
                 'SSTOT                                        001    ': "float32",
                 'TADMA                                        001    ': "float32",
                 'TADMI                                        001    ': "float32",
                 'TCAM                                         001    ': "float32",
                 'TDEG                                         001    ': "float32",
                 'TPAM                                         001    ': "float32",
                 'TPRECIP                                      001    ': "float32",
                 'TSOUT                                        001    ': "float32",
                 'TSRUN                                        001    ': "float32",
                 'WWW                                          001    ': "float32"}
     
        # read all files
        for i in range(len( modelruns)):
            runid = modelruns.iloc[i]              
            try:
                #laod source file, add ids and new column names
                tmp = pd.read_csv(self.fpath + runid["ID"] + "/macro001.csv",dtype=dtype)
                print(runid["ID"])
                tmp.columns = columns_new
                tmp["id"]=runid["ID"]
                scenario = runid["ID"].split("_")
                tmp["Tillage"] = scenario[0]
                tmp["Scenario"] = scenario[1]
                tmp["Field"] = scenario[2]
                tmp["new_date"] = tmp["Date"].apply(lambda x: pd.to_datetime(str(x)[:4] + "-" + str(x)[4:6]+ "-" + str(x)[6:8] + " " + str(x)[8:10] + ":" + str(x)[10:]))        
                tmp.set_index("id",inplace=True)
    
                #create new file if not existing, else append
                if not (os.path.isfile(self.fpath+"macrosummary.csv")):
                    tmp.to_csv(self.fpath+"macrosummary.csv")      
                else:
                    tmp.to_csv(self.fpath+"macrosummary.csv", mode='a', header=False)     
            except:
                print("ERROR in file " + runid["ID"])

    def create_fromGAP_CROP(self,years,dummy_crop = "D4_Maize"):
        runperiod = len(years) # years
        #fieldname
        #field = "f26"
        fields = pd.unique(self.helper_GAP["points_id"])
        
        
        
        crops_parfile = []
        for field in fields:
            #get current dataset
            dat = self.helper_GAP[self.helper_GAP["points_id"] == field]
            #define generic header for crop  setting
            crops_header = [[field,"AA_NCROPS",runperiod]]
            #create header
            crops_parfile += crops_header
        #    create irrigation events for each year
            for y,year in enumerate(years):                   
                d = dat[dat["Year"]==year]
 
                if len(d)>0: # crop exists in GAP    
                    crop =self.helper_crop[self.helper_crop["ID"]==d["CropGroup"].values[0]]
                    crop_template=[[field,"ATTEN",1,0.6],[field,"BETA",1,0.2],[field,"CANCAP",1,	3],[field,"CFORM",1,1.7],[field,"CRITAIR",1,5],[field,"DFORM",1,0.3],[field,"FAWC",1,0.35],[field,"HCROP",1,0.5],[field,"HMAX",1,0],[field,"IDMAX",1,230],[field,"IDMAX2",1,	0],[field,"IDSTART",1,130],[field,"IDSTART2",1,0],[field,"IHARV",1,255],[field,"IHARV2",1,	0],[field,"LAIC",1,5],[field,"LAIHAR",1,2],[field,"LAIMAX",1,5],[field,"LAIMIN",1,	0.01],[field,"NCROP",1,1],[field,"RI50",1,55],[field,"ROOTDEP",1,0.5],[field,"ROOTINIT",1,0.01],[field,"ROOTMAX",1,1],[field,"RPIN",1,67],[field,"RSMIN",1,80],[field,"RSURF",1,60],[field,"WATEN",1,1.69987449038],[field,"VPD50",1,100],[field,"ZALP",1,1.5],[field,"ZDATEMIN",1,131],[field,"ZDATEMIN2",1,0],[field,"ZHMIN",1,0.01]]
                    for param in crop_template:
                        param[2] = y+1
                        param[3] = crop[param[1]].values[0]
                else:
                    crop =self.helper_crop[self.helper_crop["ID"]==dummy_crop]
                    crop_template=[[field,"ATTEN",1,0.6],[field,"BETA",1,0.2],[field,"CANCAP",1,	3],[field,"CFORM",1,1.7],[field,"CRITAIR",1,5],[field,"DFORM",1,0.3],[field,"FAWC",1,0.35],[field,"HCROP",1,0.5],[field,"HMAX",1,0],[field,"IDMAX",1,230],[field,"IDMAX2",1,	0],[field,"IDSTART",1,130],[field,"IDSTART2",1,0],[field,"IHARV",1,255],[field,"IHARV2",1,	0],[field,"LAIC",1,5],[field,"LAIHAR",1,2],[field,"LAIMAX",1,5],[field,"LAIMIN",1,	0.01],[field,"NCROP",1,1],[field,"RI50",1,55],[field,"ROOTDEP",1,0.5],[field,"ROOTINIT",1,0.01],[field,"ROOTMAX",1,1],[field,"RPIN",1,67],[field,"RSMIN",1,80],[field,"RSURF",1,60],[field,"WATEN",1,1.69987449038],[field,"VPD50",1,100],[field,"ZALP",1,1.5],[field,"ZDATEMIN",1,131],[field,"ZDATEMIN2",1,0],[field,"ZHMIN",1,0.01]]
                    for param in crop_template:
                        param[2] = y+1
                        param[3] = crop[param[1]].values[0]
                crops_parfile += crop_template
        #create pandas data frame
        crops_parfile = [tuple(i) for i in crops_parfile]
        crops_parfile = pd.DataFrame(crops_parfile,columns=["ID","Parameter","val1","val2"])
        crops_parfile.sort_values(['ID', 'Parameter'], ascending=[True, True],inplace=True)
        crops_parfile["Parameter"][crops_parfile["Parameter"]=="AA_NCROPS"] = "NCROPS"
        crops_parfile.set_index("ID",inplace=True)
        #save new crops_parfile table to excel workbook
        book = load_workbook(os.path.join(self.fpath,self.fname + ".xlsx"))
        writer = pd.ExcelWriter(os.path.join(self.fpath,self.fname + ".xlsx"), engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        crops_parfile.to_excel(writer,"CROP")
        writer.save()

    def create_fromGAP_IRRIGATION(self,years):
        #fieldname
        #field = "f26"
        fields = pd.unique(self.helper_GAP["points_id"])
        irrigation_parfile = []
        for field in fields:
            #define generic header for irrgation setting
            irrigation_header = [[field,"IRRSAME","False"],[field,"CRITDEF",-1],[field,"IRRYEARS",len(years)]]
            #get current dataset
            dat = self.helper_GAP[self.helper_GAP["points_id"] == field]
            #get unique years
            ZFINT = 0 #TODO: tob be calculated
            #create header
            irrigation_parfile += irrigation_header
            #create irrigation events for each year
            for y,year in enumerate(years):                   
                d = dat[dat["Year"]==year]
                irrigation_event = [[field,"IRRYEAR",1],[field,"NIRRIGATIONS",1],[field,"IRRIGNO",1],[field,"IRRDAY",151],[field,"IRRSTART",12],[field,"IRREND",12.3],[field,"AMIR",0.1],[field,"CONCI",0],[field,"ZFINT",0]]
                if len(d)>0:                       
                    irrigation_event[0][2] =y+1
                    irrigation_event[3][2] = d["IRRDAY"].values[0]
                    irrigation_event[7][2] = d["ApplRate"].values[0]
                    irrigation_event[8][2] = ZFINT  
                else:
                    irrigation_event[0][2] =y+1
                    irrigation_event[1][2] = 1
                    irrigation_event[2][2] = 0
                    irrigation_event[3][2] = 0
                    irrigation_event[4][2] = 0
                    irrigation_event[5][2] = 0
                    irrigation_event[6][2] = 0
                    irrigation_event[7][2] = 0
                    irrigation_event[8][2] = 0                                  
                irrigation_parfile += irrigation_event

        #create pandas data frame
        irrigation_parfile = [tuple(i) for i in irrigation_parfile]
        irrigation_parfile = pd.DataFrame(irrigation_parfile,columns=["ID","Parameter","val1"])
        irrigation_parfile.set_index("ID",inplace=True)
        #save new IRRIGATION table to excel workbook
        book = load_workbook(os.path.join(self.fpath,self.fname + ".xlsx"))
        writer = pd.ExcelWriter(os.path.join(self.fpath,self.fname + ".xlsx"), engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        irrigation_parfile.to_excel(writer,"IRRIGATION")
        writer.save()
        return irrigation_parfile

    def create_Projects(self,ncores=2):
        # create one or more batchfiles according to number of cores
        # allocate model runs into sperate chunks for parallel processing
        modelruns = self.tables["MODELRUNS"]
        sims = [i+1 for i in range(len(modelruns))]
        nchunk = round(len(sims)/ncores)
        starts = [nchunk*i for i in range(ncores)][:len(sims)-1]
        ends = [i*nchunk+nchunk if (i*nchunk+nchunk+1)<len(sims) else sims[-1]  for i in range(ncores)][:len(sims)-1]
        names=[str(s)+"_"+str(e) for s,e in zip(starts,ends)]
        #create files
        for name,s,e in zip(names[:max(sims)],starts[:max(sims)],ends[:max(sims)]):
            batch = []
            batch.append("@echo off")
            #create .par file
            for i in range(0,len(modelruns))[s:e]:
                run = modelruns.iloc[i]
                self.makeFile(run["ID"] )
                self.create_projectfolder(self.fpath,run["ID"] )
                #create string for batchfile
                batch.append("cd " + '"' + self.fpath + "/" + run["ID"]  + "/" + '"')
                batch.append('call exeparfile.exe ' + '"' + run["ID"]  + '.par' + '"')
                batch.append('call conv.exe "MACRO001.BIN' + '"')     
            #create batch-file
            f = open(os.path.join(self.fpath,"slave_" + name + ".cmd"),"w")
            f.write("\n".join(batch))
            f.close()
            # preapre master batch-file to start slaves
            #create master batch-self
            f = open(os.path.join(self.fpath,"master_run.cmd"),"w")
            f.write("\n".join(["start slave_" + name + ".cmd" for name in names]))
            f.close()         
    
    def create_projectfolder(self,source_path,runID):
        # create target project path
        target_fpath = os.path.join(source_path,runID) 
        
        #get name of climate and rainfall file
        setup_info = self.tables["SETUP"]   
        setup_ID = self.tables["MODELRUNS"][self.tables["MODELRUNS"]["ID"]==runID]["SETUP"].values[0]
        current_setup = setup_info[setup_info["ID"]==setup_ID]
        METFILE = current_setup[current_setup["Parameter"]=="METFILE"]["val1"].values[0]
        RAINFALLFILE = current_setup[current_setup["Parameter"]=="RAINFALLFILE"]["val1"].values[0]
    
        # create directory
        if not os.path.exists(target_fpath): os.makedirs(target_fpath)
        # copy macro project file
        copyfile(os.path.join(source_path,str(runID)+".par"),os.path.join(target_fpath,str(runID)+".par")) 
        os.remove(os.path.join(source_path,str(runID)+".par")) 
        # copy executtables
        copyfile(os.path.join(source_path,"exeparfile.exe"),os.path.join(target_fpath,"exeparfile.exe"))
        copyfile(os.path.join(source_path,"Macro52Model.exe"),os.path.join(target_fpath,"Macro52Model.exe"))
        copyfile(os.path.join(source_path,"MACRO5.dll"),os.path.join(target_fpath,"MACRO5.dll"))  
        copyfile(os.path.join(source_path,"adodb.dll"),os.path.join(target_fpath,"adodb.dll")) 
        copyfile(os.path.join(source_path,METFILE),os.path.join(target_fpath,METFILE)  )
        copyfile(os.path.join(source_path,RAINFALLFILE),os.path.join(target_fpath,RAINFALLFILE))  
        copyfile(os.path.join(source_path,"conv.exe"),os.path.join(target_fpath,"conv.exe"))  
        return target_fpath   
    
    def makeFile(self,runID):
                  
        #create parameter file
        self.create_ParameterFile(runID)
        modelrun =  self.tables["MODELRUNS"][ self.tables["MODELRUNS"]["ID"]==runID]
        # replace scenario paramter
        params_to_replace = [[i.split("_")[1],i.split("_")[2],modelrun[i].values[0]] for i in modelrun.columns if i.split("_")[0]=="replace"]
        self.modify_byName(runID,params_to_replace)
        
        ###################################################################
        # calc new ZDK 
        properties = self.tables["PROPERTIES"][self.tables["PROPERTIES"]["ID"]==modelrun["PROPERTIES"].values[0]]
        Corg = properties[properties["Parameter"] == "ORGC"]["val2"].values
        solute=self.tables["SOLUTE"][self.tables["SOLUTE"]["ID"]==modelrun["SOLUTE"].values[0]]
        KOC=solute[solute["Parameter"]=="KOC"]["val1"].values[0]
        ZKD=self.calc_ZKD(Corg,KOC)
        layer = [i+1 for i,_ in enumerate(ZKD)]
        params_to_replace = [("ZKD",2,zkd,layer) for zkd,layer in zip(ZKD,layer)]
        self.modify_byName_Layer(runID,params_to_replace)
        
        ###################################################################
        #calc depth dependend solute params
        soilname=modelrun["PROPERTIES"].values[0]
        properties = self.tables["PROPERTIES"][self.tables["PROPERTIES"]["ID"]==soilname]
        thickness = properties[properties["Parameter"] == "THICKNESS"]["val2"].values
        lower_depth = np.cumsum(thickness)
        layer_center_cm = (lower_depth+(lower_depth-thickness))/2
        #calc degradation
        DT50=FREUND = modelrun["DT50"].values[0]
        DEGRADATION = self.calc_degradation_factors(layer_center_cm,DT50)
        #create list fpr replacement
        params_to_replace2=[]
        for param in  ["DEGMAL","DEGMAS","DEGMIL","DEGMIS"]:
            params_to_replace2 += [(param,2,deg,layer) for deg,layer in zip(DEGRADATION,layer)]
        #FREUND
        FREUND = modelrun["FREUND"].values[0]
        params_to_replace2 += [("FREUND",2,FREUND,l) for l in layer]
        #repalce values
        self.modify_byName_Layer(runID,params_to_replace2)
        
        ###################################################################
        # calc WATEN
        crop = modelrun["CROP"].values[0]
        croptable = self.tables["CROP"][self.tables["CROP"]["ID"]==crop] 
        fawc=croptable[croptable["Parameter"]=="FAWC"]["val2"].values[0]
        rootmax=croptable[croptable["Parameter"]=="ROOTMAX"]["val2"].values[0]         
        clay=properties[properties["Parameter"] == "CLAY"]["val2"].values
        physical = self.tables["PHYSICAL"][self.tables["PHYSICAL"]["ID"]==modelrun["PROPERTIES"].values[0]]
        bulkdensity = physical[physical["Parameter"] == "GAMMA"]["val2"].values
        silt = self.helper_soil["silt"][self.helper_soil["soil"]==soilname].values
        WATEN = []
        for c,s,bd,corg,depth in zip(clay,silt,bulkdensity,Corg,lower_depth):
            if depth<=rootmax*100:
                if depth<30: 
                    topsoil = 1 
                else: 
                    topsoil =0
             
                WATEN.append(self.calc_WATEN(fawc,c,s,bd,corg,topsoil,RESID = 0,CTEN = 10))
        WATEN = np.mean(WATEN)
        self.modify_byName(runID,[("WATEN",2,WATEN)])

    def conv_type(self,val):
        #check if date
        if type(val) == datetime.datetime:
            return "%02d/%02d/%02d"%(val.month,val.day,int(str(val.year)[2:]))
        elif pd.isnull(val):
            return ""
        else:
            #check if string
            if type(val) == str:
                return val
            else:
                #check if integer
                if val%1==0:
                    return (str(int(val)))
                else:# if not int: float
                    return str(float(val))
        
    def create_ParameterFile(self,runID):
                             
        #get infos for model run 
        modelruns = self.tables["MODELRUNS"]
        modelrun=modelruns[modelruns["ID"]==runID]
        #collect data and create string
        par = []
        for table_id in self.tablenames[1:]:   
            #get current ID of set
            set_id = modelrun[table_id].values[0]
            
            
            #get vals from table of set
            table_vals = self.tables[table_id][self.tables[table_id]["ID"]==set_id]
            #create string
            par.append("\n".join(["\t".join([self.conv_type(col) for col in row[1:] if not (pd.isnull(col))]) for row in table_vals.values]))
        s=[]
        for p,header in zip(par,self.headers_parfile):
            if header !="":
                s.append("********************************")    
                s.append(header)
                s.append(p)
            else:
                s.append(p)
        #save text file
        f = open(self.fpath+ "/" + runID+".par","w")
        f.write("\n".join(s))
        f.close()
    
    def modify_byName(self,runID,params_to_replace):
        #open original file
        f = open(self.fpath +"/" + runID+".par","r")
        dat=f.read()
        f.close()
        dat=dat.split("\n")
        dat=[i.split("\t") for i in dat]
        def setvalue(dat,param,value,column):
            for i,val in enumerate(dat):
                if val[0]==param:
                    dat[i][column]=str(value)
            return dat
        #change vlaues
        for param,column,value in params_to_replace:
            dat = setvalue(dat,param,value,int(column))
        #save file
        dat = ["\t".join(i) for i in dat]
        dat = "\n".join(dat)
        f = open(self.fpath+"/" +runID+".par","w")
        f.write(dat)
        f.close()

    def modify_byName_Layer(self,runID,params_to_replace):
        #open original file
        f = open(self.fpath+"/" +runID+".par","r")
        dat=f.read()
        f.close()
        dat=dat.split("\n")
        dat=[i.split("\t") for i in dat]
        
        def setvalue(dat,param,value,column,layer):
            for i,val in enumerate(dat):
                if val[0]==param and val[1]==str(layer):
                    dat[i][column]=str(value)
            return dat
        #change vlaues
        for param,column,value,layer in params_to_replace:
            dat = setvalue(dat,param,value,int(column),layer)
        #save file
        dat = ["\t".join(i) for i in dat]
        dat = "\n".join(dat)
        f = open(self.fpath+"/" +runID+".par","w")
        f.write(dat)
        f.close()
            
    def calc_degradation_factors(self,layer_center_cm,DT50):
        #calc degradation factor
        deg = np.log(2)/DT50
        #adjust degradation factor to depth
        def adj(d):
            adj=1
            if d<=30: adj=1
            elif d>30 and d<=60: adj=0.5
            elif d>60 and d<=100:adj=0.3
            elif d>100:adj=0
            return adj
        return [adj(c)*deg for c in layer_center_cm]
    
    def calc_ZKD(self,Corg,KOC):
        #calc ZKD
        ZKD = Corg * KOC / 100.
        return ZKD

    def calc_WATEN(self,fawc,c,s,bd,Corg,topsoil,RESID = 0,CTEN = 10):
        """
        Calculates vanGenuchten rention curve parameter n,alpha,Ksat
        fawc (double):       field water content plant stress
        c (double):         clay content (%)
        s (double):         silt content (%)
        bd (double):         buld density (g/cm³)
        Corg (double):       Organic carbon (%)
        topsoil (integer):   indicates topsoil horizon (topsoil=1,subsoil=0)
        RESID (double):      ...
        CTEN (double):       ...
        """
        #calculate organic matte rfrom Cor
        om = Corg *1.724
        #calcualte saturated water content
        theta_s = self.calc_theta_s(c,s,bd,Corg,topsoil,stones=0,stone_porosity = 10)
        # van Genuchten params
        Ksat,alpha,n,phi = self.get_vanGenuchtenMualem_HYPRES(c,s,om,topsoil,bd,None)
        #calculation of WATEN
        xmpor = theta_s / ((1 + (alpha * CTEN)** n)**(1 - (1 /n))) # water content at CTEN
        TS = theta_s/100 #fictitious sat wc
        M = 1 - 1/n #1-1/n
        T50 = xmpor/100 #xmpor/100
        WILT = theta_s / ((1 + (alpha * 15000)**n)**(1 - (1 /n))) #water content at pF 4.2
        S=(1-fawc)*(T50-WILT/100)+WILT/100 #water content corresponding to WATEN
        TE = (S-RESID/100)/(TS-RESID/100) #effective saturation at WATEN / S
        WATEN  =0.01 * 1/alpha* ((1/(TE**(1/M)) - 1))**(1/n) # [m]
        s = "theta_s %.4f vG-N %.4f vG-alpha %.4f xmpor %.4f S %.4f TE %.4f WATEN %.2f"%(theta_s,n,alpha,xmpor,S,TE,WATEN)
        return WATEN
    
    def get_vanGenuchtenMualem_HYPRES(self,c,s,om,topsoil,bd,pf):
        """
        Calculates vanGenuchten rention curve parameter n,alpha,Ksat
        
        c (double):         clay content (%)
        s (double):         silt content (%)
        bd (double):         buld density (g/cm³)
        om (double):        organic matte (%)
        topsoil (double):   indicates topsoil horizon (topsoil=1,subsoil=0)
        pf (double):        densitiy of material
        """
        d=bd
        alpha = -14.96 + 0.03135 * c + 0.0351*s + 0.646*om +15.29*d - 0.192*topsoil - 4.671*d**2 - 0.000781*c**2 - 0.00687*om**2 + 0.0449*om**(-1) + 0.0663*np.log(s) + 0.1482*np.log(om) - 0.04546*d*s - 0.4852*d*om +0.00673*topsoil*c
        n =  -25.23 - 0.02195 * c + 0.0074 * s - 0.1940 * om + 45.5 * d - 7.24 * d**2 + 0.0003658 * c**2 + 0.002885 *om**2 -12.81*d**(-1) - 0.1524 * s**(-1) - 0.01958 * om**(-1) - 0.2876*np.log(s) - 0.0709*np.log(om) - 44.6*np.log(d) - 0.02264*d*c + 0.0896*d*om + 0.00718*topsoil*c
        Ksat = 7.755 + 0.0352 * s + 0.93 * topsoil - 0.967 * d**2 - 0.000484 * c**2 - 0.000322 * s**2 + 0.001 * s**(-1) - 0.0748 * om**(-1) - 0.643*np.log(s) - 0.01398*d*c - 0.1673*d*om + 0.02986*topsoil*c -0.03305*topsoil*s
        if pf != None:
            phi = 1 - (bd / pf)
        else:
            phi = -9999
        
        alpha = np.exp(alpha)
        n = np.exp(n)+1
        Ksat = np.exp(Ksat )*24/10/100 # convert mm/h to m/day
        return (Ksat,alpha,n,phi)
    
    def calc_theta_s(self,c,s,bd,Corg,topsoil,stones=0,stone_porosity = 10):
        """
        #Woesten + adjustment for stone content
        c (double):          clay content (%)
        s (double):          silt content (%)
        bd (double):         buld density (g/cm³)
        Corg (double):       Organic carbon (%)
        topsoil (integer):   indicates topsoil horizon (topsoil=1,subsoil=0)
        stones (double):     stone content (%)
        c (double):          clay content (%)
        """
        om = Corg *1.724
        theta_s=(1 -(stones/100 * (1 - stone_porosity/100) ) ) * ( 0.7919 + (0.001691 * c) - (0.29619 * bd) - (0.000001491 * (s**2 ) ) + (0.0000821 * (om**2 )) + (0.02427/c) + ( 0.01113/s ) + (0.01472 * np.log(s)) - (0.0000733 * Corg * 1.724 * c) - (0.000619 * bd * c) - (0.001183 * bd * Corg * 1.724 ) -  (0.0001664 * topsoil * s) ) * 100
        return theta_s

    def runExecutable(self,target_fpath,runID):
        os.chdir(target_fpath)
        process = subprocess.Popen(target_fpath + "exeparfile.exe " + runID + ".par", shell=True,  stdin=None, stdout=None, stderr=None, close_fds=True)
        out, err = process.communicate()
        process.wait()

    def check_modelruns(self):
        #check all model runs
        for _,run in self.tables["MODELRUNS"].iterrows():
            #get file path
            fpath = self.fpath+run["ID"] +"/macro001.csv"
            #cehck if file exists and accessable
            if os.path.isfile(fpath) and os.access(fpath, os.R_OK):
                self.modelruns_status.append ([fpath,"OK"])
            else:
                self.modelruns_status.append ([fpath,"ERROR"])






       
    def convert_to_CMF(self,fpath_field):

    
        def make_cmf_file(d,macro_path):
            fieldname = d.index[0]
            d_cfm = pd.DataFrame()
            d_cfm["new_date"] = d["new_date"]
            d_cfm["qperc"] = d["qperc_m3h"]
            d_cfm["qsurf"] = d["qsurf_m3h"]
            d_cfm["qdrain"] = d["qdrain1st_m3h"]+d["qdrain2nd_m3h"]
            d_cfm["concgw"] = d["loadgw_mgh"] / d["qperc_m3h"]
            d_cfm["concsw"] =  d["loadsw_mgh"] / d["qsurf_m3h"]
            d_cfm["concdrain"] = (d["loaddr1_mgh"]+d["loaddr2_mgh"])/(d["qdrain1st_m3h"]+d["qdrain2nd_m3h"])
            d_cfm["fieldname"] = fieldname
            d_cfm.set_index("fieldname",inplace=True)
            #save file
            d_cfm.to_csv(macro_path +"AgriculturalFields/"+fieldname+".csv")
        
        def convert_formats(d,area):
        
            # convert flows and multiply by arare to derive total flow in m3 per hour
            d["qperc_m3h"] = d['Perc [mm/h]'].values * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
            d["qsurf_m3h"] = d['Runoff [mm/h]'].values * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
            d["qdrain1st_m3h"] =d['Drainage [mm/h]'].values  * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
            d["qdrain2nd_m3h"] =d['Drainage sec. [mm/h]'].values  * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
            # delete values lower than zero
            d["qperc_m3h"][d["qperc_m3h"]<0.0001]=0
            d["qsurf_m3h"][d["qsurf_m3h"]<0.0001]=0
            d["qdrain1st_m3h"][d["qdrain1st_m3h"]<0.0001]=0
            d["qdrain2nd_m3h"][d["qdrain2nd_m3h"]<0.0001]=0
            # convert accumalted data to rate
            d["Solutes perc rate [g/ha]"] = d["Solutes perc accumulated [mg/m2]"].diff() * 10 # convert to g/ha
            d["Solutes drainage rate [g/ha]"] = d["Solutes drainage accumulated [mg/m2]"].diff() * 10 # convert to g/ha
            d["Solutes sec drainage rate [g/ha]"] = d["Solutes sec drainage accumulated [mg/m2]"].diff() * 10 # convert to g/ha
            d["Solutes runoff rate [g/ha]"] = d["Solutes runoff accumulated [mg/m2]"].diff() * 10 # convert to g/ha
            #delete zeros
            d['Solutes perc rate [g/ha]'][d['Solutes perc rate [g/ha]']<0.0001]=0
            d['Solutes drainage rate [g/ha]'][d['Solutes drainage rate [g/ha]']<0.0001]=0
            d['Solutes sec drainage rate [g/ha]'][d['Solutes sec drainage rate [g/ha]']<0.0001]=0
            d['Solutes runoff rate [g/ha]'][d['Solutes runoff rate [g/ha]']<0.0001]=0
            # calc total loads
            d["loadgw_mgh"] = d['Solutes perc rate [g/ha]'].values * 0.1 * area
            d["loaddr1_mgh"] = d['Solutes drainage rate [g/ha]'].values * 0.1 * area
            d["loaddr2_mgh"] = d['Solutes sec drainage rate [g/ha]'].values * 0.1 * area
            d["loadsw_mgh"] = d['Solutes runoff rate [g/ha]'].values * 0.1 * area
            d["loadsum_mgh"] = d["loadgw_mgh"].values + d["loaddr1_mgh"].values + d["loaddr2_mgh"].values + d["loadsw_mgh"].values
            #create new file
            all_data = d[["fieldname","new_date","qperc_m3h","qsurf_m3h","qdrain1st_m3h","qdrain2nd_m3h","loadgw_mgh","loaddr1_mgh","loaddr2_mgh","loadsw_mgh","loadsum_mgh"]]
            all_data.set_index("fieldname",inplace=True)
            return all_data        
        #    
        
        #create new folder for field results
        if not os.path.exists(self.fpath + "AgriculturalFields"):
            os.mkdir(self.fpath + "AgriculturalFields")
        
        #read summary file
        dat = pd.read_csv(self.fpath + "macrosummary.csv")
        dat["name"]=dat["id"].apply(lambda x: x.split("_")[-1])
        
           # read fields
        field_list = pd.read_csv("N:/MOD/106520_BCS_catchment_models/mod/cal/model_runs/final/GKB2_SubB_tsMACRO/Fields.csv")
        
        #create file with all data
        alldata = pd.DataFrame()
        
        #create remaining files
        for i,field in  field_list.iterrows():
            # get  fieldsvalues of current
            d = dat[dat["name"]==field["name"]]
            # send user message
            print(field["name"],os.path.exists(self.fpath + "AgriculturalFields/"+field["name"]+".csv"))
            # check if the specific field has been modelled with MACRO
            if not len(d)>0:
                d = dat[dat["name"]==field["soil"]]
            d["fieldname"] =field["name"]
            #convert formats
            d=convert_formats(d,field["area"])
            #make cmf file
            make_cmf_file(d,self.fpath)        
        
            #append to all data
            alldata = alldata.append(d)
        
        #create sumamry of all data
        alldata.reset_index(inplace=True)
        # group data by date
        all_data_summary = alldata.groupby("new_date").sum()
        # save data
        all_data_summary[['qperc_m3h', 'qsurf_m3h', 'qdrain1st_m3h',  
                          'qdrain2nd_m3h',"loadgw_mgh","loaddr1_mgh",
                          "loaddr2_mgh","loadsw_mgh","loadsum_mgh"]].to_csv(macro.fpath + "alldata_summary.csv")


       
#    def convert_to_CMF(self,fpath_field):
#        
#        def make_cmf_file(d,macro_path):
#     
#            fieldname = d.index[0]
#            d_cfm = pd.DataFrame()
#            d_cfm["new_date"] = d["new_date"]
#            d_cfm["qperc"] = d["qperc_m3h"]
#            d_cfm["qsurf"] = d["qsurf_m3h"]
#            d_cfm["qdrain"] = d["qdrain1st_m3h"]+d["qdrain2nd_m3h"]
#            d_cfm["concgw"] = d["loadgw_mgh"] / d["qperc_m3h"]
#            d_cfm["concsw"] =  d["loadsw_mgh"] / d["qsurf_m3h"]
#            d_cfm["concdrain"] = (d["loaddr1_mgh"]+d["loaddr2_mgh"])/(d["qdrain1st_m3h"]+d["qdrain2nd_m3h"])
#            d_cfm["fieldname"] = field["name"]
#            d_cfm.set_index("fieldname",inplace=True)
#            #save file
#            d_cfm.to_csv(macro_path +"AgriculturalFields/"+fieldname+".csv")
#        
#    
#        def convert_formats(d,area):
#        
#            # convert flows and multiply by arare to derive total flow in m3 per hour
#            d["qperc_m3h"] = d['Perc [mm/h]'].values * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
#            d["qsurf_m3h"] = d['Runoff [mm/h]'].values * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
#            d["qdrain1st_m3h"] =d['Drainage [mm/h]'].values  * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
#            d["qdrain2nd_m3h"] =d['Drainage sec. [mm/h]'].values  * 0.001 * area # convert L/m2 to m3/ha and multiply by area in ha
#            # delete values lower than zero
#            d["qperc_m3h"][d["qperc_m3h"]<0.0001]=0
#            d["qsurf_m3h"][d["qsurf_m3h"]<0.0001]=0
#            d["qdrain1st_m3h"][d["qdrain1st_m3h"]<0.0001]=0
#            d["qdrain2nd_m3h"][d["qdrain2nd_m3h"]<0.0001]=0
#            # convert accumalted data to rate
#            d["Solutes perc rate [g/ha]"] = d["Solutes perc accumulated [mg/m2]"].diff() * 10 # convert to g/ha
#            d["Solutes drainage rate [g/ha]"] = d["Solutes drainage accumulated [mg/m2]"].diff() * 10 # convert to g/ha
#            d["Solutes sec drainage rate [g/ha]"] = d["Solutes sec drainage accumulated [mg/m2]"].diff() * 10 # convert to g/ha
#            d["Solutes runoff rate [g/ha]"] = d["Solutes runoff accumulated [mg/m2]"].diff() * 10 # convert to g/ha
#            #delete zeros
#            d['Solutes perc rate [g/ha]'][d['Solutes perc rate [g/ha]']<0.0001]=0
#            d['Solutes drainage rate [g/ha]'][d['Solutes drainage rate [g/ha]']<0.0001]=0
#            d['Solutes sec drainage rate [g/ha]'][d['Solutes sec drainage rate [g/ha]']<0.0001]=0
#            d['Solutes runoff rate [g/ha]'][d['Solutes runoff rate [g/ha]']<0.0001]=0
#            # calc total loads
#            d["loadgw_mgh"] = d['Solutes perc rate [g/ha]'].values * 0.1 * area
#            d["loaddr1_mgh"] = d['Solutes drainage rate [g/ha]'].values * 0.1 * area
#            d["loaddr2_mgh"] = d['Solutes sec drainage rate [g/ha]'].values * 0.1 * area
#            d["loadsw_mgh"] = d['Solutes runoff rate [g/ha]'].values * 0.1 * area
#            d["loadsum_mgh"] = d["loadgw_mgh"].values + d["loaddr1_mgh"].values + d["loaddr2_mgh"].values + d["loadsw_mgh"].values
#            #create new file
#            all_data = d[["fieldname","new_date","qperc_m3h","qsurf_m3h","qdrain1st_m3h","qdrain2nd_m3h","loadgw_mgh","loaddr1_mgh","loaddr2_mgh","loadsw_mgh","loadsum_mgh"]]
#            all_data.set_index("fieldname",inplace=True)
#            return all_data        
#    #    
#    
#        #create new folder for field results
#        if not os.path.exists(self.fpath + "AgriculturalFields"):
#            os.mkdir(self.fpath + "AgriculturalFields")
#        
#        #read summary file
#        dat = pd.read_csv(self.fpath + "macrosummary.csv")
#        dat["name"]=dat["id"].apply(lambda x: x.split("_")[-1])
#    
#       # read fields
#        field_list = pd.read_csv(fpath_field)
#        
#        alldata=pd.DataFrame()
#        
#         #create field data where no specific management has been applied
#        for _,field in field_list.iterrows():
#            d = dat[dat["name"]==field["name"]]
#            # if no specifc mdoel setuo exists use generic soil
#            if not len(d)>0:
#                d = dat[dat["name"]==field["soil"]]
#            # somer user info
#            print(field["name"],len(d))        
#            #creste cmf file
#            d = dat[dat["name"]==field["soil"]]
#            d["fieldname"] =field["name"]
#            #convert formats
#            d=convert_formats(d,field["area"])
#            alldata = alldata.append(d)
#            # make cmf file
#            make_cmf_file(d,self.fpath)   
#
#        alldata.reset_index(inplace=True)
#        alldata = alldata.groupby("new_date").sum()
#        alldata.to_csv(self.fpath+ "alldata_summary.csv")
           


              
            
if __name__ == "__main__":
    
    ################################################################################           
    ###create macro projects
    #
    ## open macro setup file     
    #macro = MACRO_pestmod("GKB2_MACRO_SCENARIOS_FINAL_KSAT","C:/MACRO_FLU_20102013")
    ##create irrigation settings from GAP
    #irrigation_parfile=macro.create_fromGAP_IRRIGATION([2010,2011,2012,2013])
    ##create crop settings from GAP
    #macro.create_fromGAP_CROP([2010,2011,2012,2013],dummy_crop = "D4_Maize")
    ##create project folders and batch-files
    #macro = MACRO_pestmod("GKB2_MACRO_SCENARIOS_FINAL_KSAT","C:/MACRO_FLU_20102013")
    #macro.create_Projects(ncores=4)
    
#    ###############################################################################
#    # create summary of results
#    
    # open macro setup file
    macro = MACRO_pestmod("GKB2_MACRO_SCENARIOS_FINAL_KSAT","n:/MOD/106520_BCS_catchment_models/mod/cal/MACRO/MACRO_FLU_20102013/")
#    
#    #check runs
#    macro.check_modelruns()
#    print(macro.modelruns_status)

#     create data summary
#    macro.create_results_summary()

    # convert to cmf
#    macro.convert_to_CMF("N:/MOD/106520_BCS_catchment_models/mod/cal/model_runs/final/GKB2_SubB_tsMACRO/Fields.csv")







    





















#    #create one summary time series of all fields
#    fpath = "n:/MOD/106520_BCS_catchment_models/mod/cal/MACRO/MACRO_METR_20102013/"
#    files = os.listdir(fpath)
#    res = pd.read_csv(fpath + files[0])
#    for f in files[1:]:
#        print(f)
#        res = res.append(pd.read_csv(fpath + files[0]))

##    
#    res_summary = res.groupby("new_date").sum()
#    res_summary.reset_index(inplace=True)
#    res_summary.set_index("new_date",inplace=True)
#    res_summary.to_csv("n:/MOD/106520_BCS_catchment_models/mod/cal/MACRO/MACRO_METR_20102013/alldata_summary.csv")
#
#

    #################################################################################
    # make plot of simulated and observed data

#    #load all data summary
#    all_data_summary = pd.read_csv(macro.fpath + "alldata_summary.csv")
##    
#    # simulated flow
#    flow = all_data_summary[['new_date','qperc_m3h', 'qsurf_m3h']]
#    flow["qdrain_m3h"] = all_data_summary["qdrain1st_m3h"] + all_data_summary["qdrain2nd_m3h"]
#    flow["qsum_m3h"] = flow["qperc_m3h"] + flow["qsurf_m3h"] +flow["qdrain_m3h"]
#    flow["new_date"] = pd.to_datetime(flow["new_date"])
#    flow.set_index("new_date",inplace=True)
#    flow = flow.resample("60min").mean() # results of MACRO are given per half hour, e.g. 12:30; convert to 12:00
#    
#    #simulated load
#    load = all_data_summary[['new_date','loadgw_mgh', 'loadsw_mgh',  'loadsum_mgh']]
#    load["loaddr_mgh"] = all_data_summary["loaddr1_mgh"] + all_data_summary["loaddr2_mgh"]
#    load["new_date"] = pd.to_datetime(load["new_date"])
#    load.set_index("new_date",inplace=True)
#    load = load.resample("60min").mean() # results of MACRO are given per half hour, e.g. 12:30; convert to 12:00
#    
#    #simulated concentration
#    conc = pd.DataFrame()
#    conc=pd.DataFrame()
#    conc["conc_mgm3"] = load["loadsum_mgh"] / flow["qsum_m3h"]
#    
#    #calculate flow in m3/sec
#    flow = flow/3600#convert m3/h to m3/sec
#    flow.columns = ['qperc_m3sec', 'qsurf_m3sec',  'qdrain_m3sec','qsum_m3sec']
#    
##    #save files
#    flow.to_csv(macro.fpath+"flow.csv")
#    load.to_csv("load.csv")
#    conc.to_csv("conc.csv")
        
    #observed data
#    xls = pd.ExcelFile(r"n:\MOD\106520_BCS_catchment_models\mod\cal\MACRO\psm_all_daily_final.xlsx")
#    obs = xls.parse("obs")
#    obs.set_index("Date",inplace=True)
#    rain = xls.parse("rainfall")
#    rain["Date"] = pd.to_datetime(rain["Date"])
#    rain.set_index("Date",inplace=True)
#    #applications
#    appl = xls.parse("appl")
#    appl["Date"] = pd.to_datetime(appl["Date"])
#    appl.set_index("Date",inplace=True)


#    start = "2010-05-17"
#    end = "2013-12-30"
#    name = "all"
#    
#    
##    starts=["2010-05-17","2010-05-17","2011-01-01","2012-01-01","2013-01-01"]
##    ends=["2013-12-30","2010-12-30","2011-12-30","2012-12-30","2013-12-30"]
##    names=["all","2010","2011","2012","2013"]
##    for start,end,name in zip(starts,ends,names):
#    print(name)
#    fig = plt.figure(figsize=(10,15))
#    font = {'family' : 'normal',
#            'weight' : 'normal',
#            'size'   : 10}
#    matplotlib.rc('font', **font)
    
#    #select rainfall data in time series
#    r = rain[(rain.index >= pd.Timestamp(start)) & (rain.index <= pd.Timestamp(end))]
#    #plot rainfall
#    ax1 = fig.add_axes([0.1,0.85,0.8,0.1]) # x,y, lenght, height
#    ax1.bar(r.index, r['Precipitation'].values, color="k",edgecolor=None,align='center', alpha=1)
#    ax1.invert_yaxis()
#    ax1.xaxis.tick_top()
#    ax1.xaxis.set_ticks_position('both') # THIS IS THE ONLY CHANGE
#    ax1.xaxis_date()
#    ax1.set_ylabel('Rain [mm h$^{-1}$]')
#    ax1.grid(True)
#    ax1.spines['bottom'].set_color('none')
#    ax1.xaxis.set_ticks_position('top')
#    ax1.set_yticks([20,40,60,80])
#    ax1.set_yticklabels(["20","40","60",""])
#    ax1.text(.015,.05,'(a)', horizontalalignment='left', transform=ax1.transAxes)
#    
#    #select flow data
#    ax2 = fig.add_axes([0.1,0.6,0.8,0.24]) # x,y, lenght, height
#    #make plot of observed stream flow
#    o = obs[(obs.index >= pd.Timestamp(start)) & (obs.index <= pd.Timestamp(end))]
#    ax2.plot(o.index,o['Flow [m3/sec]'],linestyle="-",color="k",linewidth=3,alpha=.25)
#    ax2.grid(True)
#    ax2.spines['top'].set_color('none')
#    ax2.xaxis.set_ticks_position('bottom')
#    ax2.set_ylabel('Stream flow [m$^3$ sec$^{-1}$]')
#    #make plot of simulated flow
#    f = flow[(flow.index >= pd.Timestamp(start)) & (flow.index <= pd.Timestamp(end))]
#    ax2.stackplot(f.index,np.row_stack((f['qperc_m3sec'], f['qsurf_m3sec'], f['qdrain_m3sec'])),alpha=1,linewidth=0,colors=["forestgreen","steelblue","orange"])
#    ax2.set_ylim(0,1.2)
#    ax2.set_xlabel("")
#    ax2.set_xticklabels([""])
#    ax2.text(.015,.97,'(b)', horizontalalignment='left', transform=ax2.transAxes)
#    
#    #make plot of conc
#    ax3 = fig.add_axes([0.1,0.35,0.8,0.24])
#    ax3.plot(o.index,o["FLU conc [mg/m3]"],linestyle="-",color="k",linewidth=3,alpha=.25)
#    c = conc[(conc.index >= pd.Timestamp(start)) & (conc.index <= pd.Timestamp(end))]
##    ax3.plot(c.index,c['conc_mgm3'],linestyle="-",color="b",linewidth=1)
#    ax3.grid(True)
#    ax3.spines['top'].set_color('none')
#    ax3.xaxis.set_ticks_position('bottom')
#    ax3.set_ylabel('Concentration [$\mu$g L$^{-1}$]')
#    ax3.set_yticks(np.arange(0,4,0.5))
#    ax3.set_xlabel("")
#    ax3.set_xticklabels([""])
#    ax3_1 = ax3.twinx()
#    appl_period = appl[(appl.index >= pd.Timestamp(start)) & (appl.index <= pd.Timestamp(end))]
#    ax3_1.plot(appl_period.index,appl_period['Appl [g/ha]'],linewidth=0,marker="x",color="k",markersize=5)
#    ax3_1.set_ylim(0,650)
#    ax3_1.set_ylabel("Application rate [g a.s. ha$^{-1}]$")
#    
#    if name=="all":
#        ax3.text(.015,.97,'(d)', horizontalalignment='center', transform=ax3.transAxes)
#    else:
#        ax3.text(.015,.97,'(3)', horizontalalignment='center', transform=ax3.transAxes)
#    #make plot of conc
#    ax4 = fig.add_axes([0.1,0.1,0.8,0.24])
#    ax4.plot(o.index,o['FLU load  [mg]'],linestyle="-",color="k",linewidth=3,alpha=.25)
#    l= load[(load.index >= pd.Timestamp(start)) & (load.index <= pd.Timestamp(end))]
#    ax4.plot(l.index,l['loadsum_mgh'],linestyle="-",color="b",linewidth=1)
#    ax4.grid(True)
#    ax4.spines['top'].set_color('none')
#    ax4.xaxis.set_ticks_position('bottom')
#    ax4.set_ylabel('Load [mg]')
#    ax4.set_ylim(0,1000)
#    ax4.set_yticks(np.arange(0,1000,200))
#    if name=="all":
#        ax4.text(.015,.97,'(e)', horizontalalignment='center', transform=ax4.transAxes)
#    else:
#        ax4.text(.015,.97,'(d)', horizontalalignment='center', transform=ax4.transAxes)
#    #make legend
#    ax5  = fig.add_axes([0.5,-0.01,0.1,0.1])
#    ax5.axis('off')
#    legend_rainfall = mlines.Line2D([],[],color='k', label='Rainfall',alpha=1,linewidth=1)
#    legend_obsflow = mlines.Line2D([],[], label='Observed flow',color="0.75",alpha=.25,linewidth=.5)
#    legend_qperc_m3sec = mpatches.Patch(color='forestgreen', label='Simulated groundwater',alpha=.75,linewidth=0)
#    legend_qsurf_m3sec = mpatches.Patch(color='steelblue', label='Simulated surface',alpha=.75,linewidth=0)
#    legend_qdrain_m3sec = mpatches.Patch(color='orange', label='Simulated drainage',alpha=.75,linewidth=0)
#    legend_obsconc = mlines.Line2D([],[], label='Observed',color="k",alpha=.25,linewidth=3)
#    legend_simconc = mlines.Line2D([],[], label='Simulated substance',color="b",alpha=1,linewidth=1)
#    legend_appl = mlines.Line2D([],[], label='Application rate',linewidth=0,marker="x",color="k",markersize=5)
#    plt.legend(handles=[legend_qperc_m3sec,legend_qsurf_m3sec,legend_qdrain_m3sec,legend_rainfall,legend_obsconc,legend_simconc,legend_appl],ncol=3, loc="center", fontsize=10.,fancybox=True, framealpha=0.5)
#    
#    if name=="all":
#        #plot barchart of load
#        ax5  = fig.add_axes([0.75,0.21,0.1,0.1])
#        #ax5.axis('off')
#        load["year"] = load.index.year
#        load_sums = load[['year','loadgw_mgh', 'loadsw_mgh','loaddr_mgh']].groupby(["year"]).sum() / 1000
#        load_sums.plot.barh(ax=ax5, stacked=True,colors=["forestgreen","steelblue","orange"],edgecolor="",legend=False, align='center')
#        #plot observed
#        obs["year"] = obs.index.year
#        obs_sum = obs[['year','FLU load  [mg]']].groupby(["year"]).sum() / 1000
#        obs_sum.plot.barh(ax=ax5,stacked=True,colors=["k"],edgecolor="",width=.1,alpha=0.5,legend=None, align='center')
#        ax5.set_ylabel("")
#        ax5.set_xlabel("[g]")
#        ax5.set_xticks(range(0,140,20))
#        ax5.set_xticklabels(range(0,140,20),rotation=270)
#        ax5.text(.5,1.05,'(f)', horizontalalignment='center', transform=ax5.transAxes)
#        
#        #plot barchart of load
#        ax6  = fig.add_axes([0.75,0.8,0.1,0.1])
#        #ax5.axis('off')
#        flow["year"] = flow.index.year
#        flow_sums = flow[['year','qperc_m3sec', 'qsurf_m3sec','qdrain_m3sec']].groupby(["year"]).sum()*3600 #m3/h
#        flow_sums = flow_sums / 350 *0.1 #m3/ha to mm
#        flow_sums.plot.barh(ax=ax6, stacked=True,colors=["forestgreen","steelblue","orange"],edgecolor="",legend=False, align='center')
#        #plot observed
#        obs["year"] = obs.index.year
#        obs_sum = obs[['year','Flow [m3/sec]']].groupby(["year"]).sum() *3600 #m3/h
#        obs_sum = obs_sum / 350 *0.1 #m3/ha to mm
#        obs_sum.plot.barh(ax=ax6,stacked=True,colors=["k"],edgecolor="",width=.1,alpha=0.5,legend=None, align='center')
#        ax6.set_ylabel("")
#        ax6.set_xlabel("[mm]")
#        ax6.set_xticklabels(range(0,400,50),rotation=270)
#        ax6.text(.5,1.05,'(c)', horizontalalignment='center', transform=ax6.transAxes)
#        
#    #save figure
#    fig.savefig(macro.fpath +name + ".png",dpi=300,transparent=True) 
#    
#
#
#
#
#
#


























################################################################################
## cals stats
#
#def stats_calc_stats(dat_observed,dat_simulated):
#    #clc r2
#    r_squared = np.corrcoef(dat_observed,dat_simulated)[0][1]**2
#
#    #calc rmse
#    def calc_rmse(obs, sim):
#        return np.sqrt(((obs - sim) ** 2).mean())
#    #calc MAPE
#    def mean_absolute_percentage_error(y_true, y_pred): 
#        y_true, y_pred = np.array(y_true), np.array(y_pred)
#        return np.mean(np.abs((y_true - y_pred) / y_true)) * 100
#    #calc NSE
#    def calc_NSE(obs,sim):
#        avg_obs = np.mean(obs)
#        return  1 - (sum((obs-sim)**2)/sum((obs-avg_obs)**2))
#    NSE = calc_NSE(dat_observed,dat_simulated)
#    #plot data
#    return r_squared,NSE

#def getDJF(dat,year):
#    dat1 = dat[(dat.index>=pd.Timestamp(year+"-12-01")) & ( dat.index<=pd.Timestamp(year+"-12-30"))]
#    dat2 = dat[ (dat.index>=pd.Timestamp(year+"-01-01")) & ( dat.index<=pd.Timestamp(year+"-02-28"))]
#    return dat1.append(dat2)
#
#def getMAM(dat,year):
#    dat1 = dat[(dat.index>=pd.Timestamp(year+"-03-01")) & ( dat.index<=pd.Timestamp(year+"-05-31"))]
#    return dat1
#
#def getJJA(dat,year):
#    dat1 = dat[(dat.index>=pd.Timestamp(year+"-06-01")) & ( dat.index<=pd.Timestamp(year+"-08-30"))]
#    return dat1
#
#def getSON(dat,year):
#    dat1 = dat[(dat.index>=pd.Timestamp(year+"-09-01")) & ( dat.index<=pd.Timestamp(year+"-11-30"))]
#    return dat1


#
#year = "2010"
#period="JJA"
#sim_merge = getJJA(res,year)
#obs_merge = getJJA(obs,year)
#r_squared,NSE = stats_calc_stats(obs_merge["Flow"],sim_merge["qsum"])
#print(year,period,r_squared,NSE)
#
#period="SON"
#sim_merge = getSON(res,year)
#obs_merge = getSON(obs,year)
#r_squared,NSE = stats_calc_stats(obs_merge["Flow"],sim_merge["qsum"])
#print(year,period,r_squared,NSE)
#    
#for year in ["2011","2012","2013"]:
#    period="DJF"
#    sim_merge = getDJF(res,year)
#    obs_merge = getDJF(obs,year)
#    r_squared,NSE = stats_calc_stats(obs_merge["Flow"],sim_merge["qsum"])
#    print(year,period,r_squared,NSE)
#
#    period="MAM"
#    sim_merge = getMAM(res,year)
#    obs_merge = getMAM(obs,year)
#    r_squared,NSE = stats_calc_stats(obs_merge["Flow"],sim_merge["qsum"])
#    print(year,period,r_squared,NSE)
#    
#    period="JJA"
#    sim_merge = getJJA(res,year)
#    obs_merge = getJJA(obs,year)
#    r_squared,NSE = stats_calc_stats(obs_merge["Flow"],sim_merge["qsum"])
#    print(year,period,r_squared,NSE)
#    
#    period="SON"
#    sim_merge = getSON(res,year)
#    obs_merge = getSON(obs,year)
#    r_squared,NSE = stats_calc_stats(obs_merge["Flow"],sim_merge["qsum"])
#    print(year,period,r_squared,NSE)



